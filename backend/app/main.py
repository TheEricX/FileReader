import os
import json
import uuid
import base64
from datetime import datetime
import pandas as pd
from typing import Dict, Set
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
import uvicorn
from dotenv import load_dotenv

from .excel_utils import ExcelUtils
from .agent import ExcelAgent
from .pdf_agent import PdfAgent
from .doc_agent import DocAgent
from .gemini_agent import GeminiAgent
from .pdf_utils import extract_pdf_text
from .doc_utils import extract_docx_text
from .persistence import (
    init_db,
    save_upload,
    save_session,
    update_session_message_history,
    get_upload,
    get_session,
    list_uploads,
    delete_upload,
    delete_session
)

load_dotenv()

app = FastAPI(title="Excel Agent API")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with specific origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global variables
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Store active connections and their associated Excel files
active_connections: Dict[str, Dict] = {}
active_pdf_connections: Dict[str, Dict] = {}
active_doc_connections: Dict[str, Dict] = {}
active_gemini_connections: Dict[str, Dict] = {}
canceled_requests: Dict[str, Set[str]] = {}

# Initialize the Excel agent
excel_agent = ExcelAgent()
pdf_agent = PdfAgent()
doc_agent = DocAgent()
gemini_agent = GeminiAgent()

class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket
    
    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]
    
    async def send_message(self, client_id: str, message: str):
        if client_id in self.active_connections:
            await self.active_connections[client_id].send_text(message)
    
    async def broadcast(self, message: str):
        for connection in self.active_connections.values():
            await connection.send_text(message)

manager = ConnectionManager()


def _mark_canceled(client_id: str, request_id: str) -> None:
    if not request_id:
        return
    if client_id not in canceled_requests:
        canceled_requests[client_id] = set()
    canceled_requests[client_id].add(request_id)


def _is_canceled(client_id: str, request_id: str) -> bool:
    return bool(request_id) and request_id in canceled_requests.get(client_id, set())


def _build_excel_system_message(file_path: str) -> str:
    file_extension = os.path.splitext(file_path)[1].lower()
    excel_utils = ExcelUtils(file_path)
    df = excel_utils.get_dataframe()
    df_head_str = df.head().to_string()
    num_rows, num_cols = df.shape
    display_rows = max(num_rows, 1)
    display_cols = max(num_cols, 1)
    from openpyxl.utils import get_column_letter
    data_range = f"A1:{get_column_letter(display_cols)}{display_rows}"
    return f"""You are a smart spreadsheet assistant. You have access to functions. Always use them when asked to read/update spreadsheet content.
        After inserting a row or column, always re-evaluate the sheet before updating it.
        Make sure the inserted index exists before trying to write to it.
        When inserting a total row, always find the last filled row index using tools. Never hardcode the index.
        Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.

        Here's a preview of top 5 rows of the spreadsheet data structure it might contains header if not you can figure it out based on data:
        {df_head_str}

        The spreadsheet contains data in the range {data_range} ({num_rows} rows × {num_cols} columns).
        File format: {file_extension[1:].upper()}"""


def _build_pdf_system_message(file_path: str, filename: str) -> str:
    pdf_text = extract_pdf_text(file_path)
    return (
        "You are a helpful PDF analysis assistant. "
        "Answer questions using the provided PDF content. "
        "If the answer is not in the PDF, say you could not find it. "
        "Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.\n\n"
        f"PDF filename: {filename}\n\n"
        f"PDF content (truncated):\n{pdf_text}"
    )


def _build_doc_system_message(filename: str, doc_text: str) -> str:
    return (
        "You are a helpful document analysis assistant. "
        "Answer questions using the provided document content. "
        "If the answer is not in the document, say you could not find it. "
        "Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.\n\n"
        f"Document filename: {filename}\n\n"
        f"Document content:\n{doc_text}"
    )


def _gemini_session_key(client_id: str) -> str:
    return f"gemini:{client_id}"


def _build_gemini_excel_system_message(file_path: str) -> str:
    file_extension = os.path.splitext(file_path)[1].lower()
    excel_utils = ExcelUtils(file_path)
    df = excel_utils.get_dataframe()
    max_full_size_mb = 1
    file_size_bytes = os.path.getsize(file_path)
    file_size_mb = file_size_bytes / (1024 * 1024)
    include_full_sheet = file_size_mb <= max_full_size_mb
    if include_full_sheet:
        preview_rows = len(df)
        df_head_str = df.to_string()
    else:
        preview_rows = min(200, len(df))
        df_head_str = df.head(preview_rows).to_string()
    num_rows, num_cols = df.shape
    display_rows = max(num_rows, 1)
    display_cols = max(num_cols, 1)
    from openpyxl.utils import get_column_letter
    data_range = f"A1:{get_column_letter(display_cols)}{display_rows}"
    return (
        "You are a helpful spreadsheet analysis assistant. "
        "You can call tools to read the spreadsheet when needed. "
        "If the answer is not in the data, say you could not find it. "
        "Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.\n\n"
        "Tool usage:\n"
        "- To call a tool, respond with ONLY a JSON object (no extra text) like:\n"
        '  {"tool":"read_excel_range","args":{"range":"A1:D20"}}\n'
        "Available tools:\n"
        "- read_excel_range(range) -> CSV text for the range (e.g., A1:D20)\n"
        "- read_cell(cell) or read_cell(row, col) (row/col are 0-based)\n"
        "- read_sheet_metadata()\n"
        "- get_column_values(index) (0-based column index)\n"
        "- filter_rows(col_index, value) (0-based column index)\n"
        "- get_last_filled_row_index()\n\n"
        f"{'Spreadsheet full content' if include_full_sheet else f'Spreadsheet preview (top {preview_rows} rows)'}:\n{df_head_str}\n\n"
        f"Spreadsheet range: {data_range} ({num_rows} rows × {num_cols} columns)\n"
        f"File format: {file_extension[1:].upper()}\n"
        f"File size: {file_size_mb:.2f} MB"
    )


def _build_gemini_pdf_system_message(file_path: str, filename: str) -> str:
    pdf_text = extract_pdf_text(file_path)
    return (
        "You are a helpful PDF analysis assistant. "
        "Answer questions using the provided PDF content. "
        "If the answer is not in the PDF, say you could not find it. "
        "Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.\n\n"
        f"PDF filename: {filename}\n\n"
        f"PDF content (truncated):\n{pdf_text}"
    )


def _build_gemini_doc_system_message(filename: str, doc_text: str) -> str:
    return (
        "You are a helpful document analysis assistant. "
        "Answer questions using the provided document content. "
        "If the answer is not in the document, say you could not find it. "
        "Format responses clearly with short sections, blank lines between sections, and bullet points when listing items.\n\n"
        f"Document filename: {filename}\n\n"
        f"Document content:\n{doc_text}"
    )


def _persist_message_history(client_id: str, session_type: str, message_history: list) -> None:
    updated = update_session_message_history(client_id, message_history)
    if not updated:
        save_session(client_id, session_type, message_history)


def _persist_gemini_message_history(client_id: str, message_history: list) -> None:
    session_key = _gemini_session_key(client_id)
    updated = update_session_message_history(session_key, message_history)
    if not updated:
        save_session(session_key, "gemini", message_history)


def _restore_excel_session(client_id: str) -> bool:
    upload = get_upload(client_id)
    if not upload or upload["type"] != "spreadsheet":
        return False
    file_path = upload["file_path"]
    if not os.path.exists(file_path):
        return False
    session = get_session(client_id)
    if session and session["message_history"]:
        message_history = session["message_history"]
    else:
        system_message = _build_excel_system_message(file_path)
        message_history = [{"role": "system", "content": system_message}]
        save_session(client_id, "spreadsheet", message_history)
    excel_utils = ExcelUtils(file_path)
    active_connections[client_id] = {
        "file_path": file_path,
        "excel_utils": excel_utils,
        "message_history": message_history
    }
    return True


def _restore_pdf_session(client_id: str) -> bool:
    upload = get_upload(client_id)
    if not upload or upload["type"] != "pdf":
        return False
    file_path = upload["file_path"]
    filename = upload.get("filename") or os.path.basename(file_path)
    if not os.path.exists(file_path):
        return False
    session = get_session(client_id)
    if session and session["message_history"]:
        message_history = session["message_history"]
    else:
        system_message = _build_pdf_system_message(file_path, filename)
        message_history = [{"role": "system", "content": system_message}]
        save_session(client_id, "pdf", message_history)
    active_pdf_connections[client_id] = {
        "file_path": file_path,
        "filename": filename,
        "message_history": message_history
    }
    return True


def _restore_doc_session(client_id: str) -> bool:
    upload = get_upload(client_id)
    if not upload or upload["type"] != "doc":
        return False
    file_path = upload["file_path"]
    filename = upload.get("filename") or os.path.basename(file_path)
    if not os.path.exists(file_path):
        return False
    session = get_session(client_id)
    if session and session["message_history"]:
        message_history = session["message_history"]
    else:
        doc_text = extract_docx_text(file_path)
        system_message = _build_doc_system_message(filename, doc_text)
        message_history = [{"role": "system", "content": system_message}]
        save_session(client_id, "doc", message_history)
    active_doc_connections[client_id] = {
        "file_path": file_path,
        "filename": filename,
        "doc_text": extract_docx_text(file_path),
        "message_history": message_history
    }
    return True


def _restore_gemini_session(client_id: str) -> bool:
    upload = get_upload(client_id)
    if not upload:
        return False
    file_path = upload["file_path"]
    if not os.path.exists(file_path):
        return False
    filename = upload.get("filename") or os.path.basename(file_path)

    session_key = _gemini_session_key(client_id)
    session = get_session(session_key)
    if session and session["message_history"]:
        message_history = session["message_history"]
    else:
        if upload["type"] == "pdf":
            system_message = _build_gemini_pdf_system_message(file_path, filename)
        elif upload["type"] == "doc":
            doc_text = extract_docx_text(file_path)
            system_message = _build_gemini_doc_system_message(filename, doc_text)
        else:
            system_message = _build_gemini_excel_system_message(file_path)
        message_history = [{"role": "system", "content": system_message}]
        save_session(session_key, "gemini", message_history)

    connection = {
        "file_path": file_path,
        "filename": filename,
        "type": upload["type"],
        "message_history": message_history
    }
    if upload["type"] == "doc":
        connection["doc_text"] = extract_docx_text(file_path)
    if upload["type"] == "spreadsheet":
        connection["excel_utils"] = ExcelUtils(file_path)
    active_gemini_connections[client_id] = connection
    return True


def _bootstrap_upload_records() -> None:
    supported_formats = {
        ".xlsx", ".xls", ".csv", ".tsv", ".ods", ".fods", ".xlsm", ".xltx", ".xltm", ".pdf", ".docx"
    }
    for filename in os.listdir(UPLOAD_DIR):
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.isfile(file_path):
            continue
        _, ext = os.path.splitext(filename)
        ext = ext.lower()
        if ext not in supported_formats:
            continue
        client_id = os.path.splitext(filename)[0]
        if get_upload(client_id):
            continue
        if ext == ".pdf":
            upload_type = "pdf"
        elif ext == ".docx":
            upload_type = "doc"
        else:
            upload_type = "spreadsheet"
        uploaded_at = datetime.utcfromtimestamp(os.path.getmtime(file_path)).isoformat()
        save_upload(client_id, upload_type, file_path, filename, uploaded_at)


def _get_or_restore_session_messages(client_id: str) -> list:
    session = get_session(client_id)
    if session and session["message_history"]:
        return session["message_history"]
    upload = get_upload(client_id)
    if not upload:
        return []
    if upload["type"] == "pdf":
        if not _restore_pdf_session(client_id):
            return []
    elif upload["type"] == "doc":
        if not _restore_doc_session(client_id):
            return []
    else:
        if not _restore_excel_session(client_id):
            return []
    session = get_session(client_id)
    if not session:
        return []
    return session.get("message_history", [])


def _get_or_restore_gemini_session_messages(client_id: str) -> list:
    session_key = _gemini_session_key(client_id)
    session = get_session(session_key)
    if session and session["message_history"]:
        return session["message_history"]
    if not _restore_gemini_session(client_id):
        return []
    session = get_session(session_key)
    if not session:
        return []
    return session.get("message_history", [])


def _reset_session_history(client_id: str, session_type: str) -> None:
    if session_type == "spreadsheet":
        if client_id in active_connections:
            file_path = active_connections[client_id]["file_path"]
            system_message = _build_excel_system_message(file_path)
            active_connections[client_id]["message_history"] = [
                {"role": "system", "content": system_message}
            ]
            save_session(client_id, "spreadsheet", active_connections[client_id]["message_history"])
            return
        upload = get_upload(client_id)
        if upload and os.path.exists(upload["file_path"]):
            system_message = _build_excel_system_message(upload["file_path"])
            save_session(client_id, "spreadsheet", [{"role": "system", "content": system_message}])
    elif session_type == "pdf":
        if client_id in active_pdf_connections:
            file_path = active_pdf_connections[client_id]["file_path"]
            filename = active_pdf_connections[client_id].get("filename") or os.path.basename(file_path)
            system_message = _build_pdf_system_message(file_path, filename)
            active_pdf_connections[client_id]["message_history"] = [
                {"role": "system", "content": system_message}
            ]
            save_session(client_id, "pdf", active_pdf_connections[client_id]["message_history"])
            return
        upload = get_upload(client_id)
        if upload and os.path.exists(upload["file_path"]):
            filename = upload.get("filename") or os.path.basename(upload["file_path"])
            system_message = _build_pdf_system_message(upload["file_path"], filename)
            save_session(client_id, "pdf", [{"role": "system", "content": system_message}])

@app.get("/")
async def root():
    return {"message": "Excel Agent API is running"}


@app.get("/uploads")
async def get_uploads():
    uploads = list_uploads()
    result = []
    for upload in uploads:
        result.append({
            "client_id": upload["client_id"],
            "type": upload["type"],
            "filename": upload.get("filename"),
            "file_url": f"/pdf/{upload['client_id']}/file" if upload["type"] == "pdf" else None,
            "uploaded_at": upload["uploaded_at"]
        })
    return {"uploads": result}


@app.get("/sessions/{client_id}")
async def get_session_messages(client_id: str):
    messages = _get_or_restore_session_messages(client_id)
    if not messages:
        return JSONResponse(status_code=404, content={"error": "Session not found"})
    return {"messages": messages}


@app.delete("/sessions/{client_id}")
async def delete_session_messages(client_id: str):
    session = get_session(client_id)
    if not session:
        return JSONResponse(status_code=404, content={"error": "Session not found"})
    delete_session(client_id)
    _reset_session_history(client_id, session["type"])
    return {"status": "cleared", "client_id": client_id}


@app.get("/sessions/gemini/{client_id}")
async def get_gemini_session_messages(client_id: str):
    messages = _get_or_restore_gemini_session_messages(client_id)
    if not messages:
        return JSONResponse(status_code=404, content={"error": "Session not found"})
    return {"messages": messages}


@app.delete("/sessions/gemini/{client_id}")
async def delete_gemini_session_messages(client_id: str):
    session_key = _gemini_session_key(client_id)
    session = get_session(session_key)
    if not session:
        return JSONResponse(status_code=404, content={"error": "Session not found"})
    delete_session(session_key)
    if client_id in active_gemini_connections:
        del active_gemini_connections[client_id]
    manager.disconnect(session_key)
    return {"status": "cleared", "client_id": client_id}

@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Upload a spreadsheet file and return a client ID for WebSocket connection"""
    print(f"Received file upload: {file.filename}")
    
    try:
        # Check file extension
        file_extension = os.path.splitext(file.filename)[1].lower()
        supported_formats = ['.xlsx', '.xls', '.csv', '.tsv', '.ods', '.fods', '.xlsm', '.xltx', '.xltm']
        
        if file_extension not in supported_formats:
            return JSONResponse(
                status_code=400,
                content={"error": f"Unsupported file format. Supported formats: {', '.join(supported_formats)}"}
            )
        
        # Generate a unique client ID
        client_id = str(uuid.uuid4())
        print(f"Generated client ID: {client_id}")
        
        # Save the uploaded file with original extension
        file_path = os.path.join(UPLOAD_DIR, f"{client_id}{file_extension}")
        print(f"Saving file to: {file_path}")
        
        with open(file_path, "wb") as buffer:
            content = await file.read()
            print(f"Read {len(content)} bytes from file")
            if not content:
                return JSONResponse(
                    status_code=400,
                    content={"error": "Uploaded file is empty."}
                )
            buffer.write(content)
        
        # Initialize Excel utils for this client
        print("Initializing Excel utils")
        excel_utils = ExcelUtils(file_path)

        # Get DataFrame head and data range information
        try:
            system_message = _build_excel_system_message(file_path)
        except Exception as e:
            error_msg = f"Failed to read spreadsheet: {str(e)}"
            print(error_msg)
            return JSONResponse(status_code=400, content={"error": error_msg})
        active_connections[client_id] = {
            "file_path": file_path,
            "excel_utils": excel_utils,
            "message_history": [
                {"role": "system", "content": system_message}
            ]
        }
        save_upload(client_id, "spreadsheet", file_path, file.filename)
        save_session(client_id, "spreadsheet", active_connections[client_id]["message_history"])
        
        print("File upload successful")
        return {"client_id": client_id}
    except Exception as e:
        error_msg = f"Error in upload_excel: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.post("/upload/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """Upload a PDF file and return a client ID for WebSocket connection"""
    print(f"Received PDF upload: {file.filename}")

    try:
        file_extension = os.path.splitext(file.filename)[1].lower()
        if file_extension != ".pdf":
            return JSONResponse(
                status_code=400,
                content={"error": "Unsupported file format. Only .pdf files are supported."}
            )

        client_id = str(uuid.uuid4())
        print(f"Generated PDF client ID: {client_id}")

        file_path = os.path.join(UPLOAD_DIR, f"{client_id}{file_extension}")
        print(f"Saving PDF to: {file_path}")

        with open(file_path, "wb") as buffer:
            content = await file.read()
            print(f"Read {len(content)} bytes from PDF")
            if not content:
                return JSONResponse(
                    status_code=400,
                    content={"error": "Uploaded PDF is empty."}
                )
            buffer.write(content)

        pdf_text = extract_pdf_text(file_path)
        if not pdf_text.strip():
            return JSONResponse(
                status_code=400,
                content={"error": "Failed to extract text from the PDF."}
            )
        system_message = _build_pdf_system_message(file_path, file.filename)

        active_pdf_connections[client_id] = {
            "file_path": file_path,
            "filename": file.filename,
            "message_history": [
                {"role": "system", "content": system_message}
            ]
        }
        save_upload(client_id, "pdf", file_path, file.filename)
        save_session(client_id, "pdf", active_pdf_connections[client_id]["message_history"])

        print("PDF upload successful")
        return {
            "client_id": client_id,
            "file_url": f"/pdf/{client_id}/file",
            "filename": file.filename
        }
    except Exception as e:
        error_msg = f"Error in upload_pdf: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})


@app.post("/upload/doc")
async def upload_doc(file: UploadFile = File(...)):
    """Upload a DOCX file and return a client ID for WebSocket connection"""
    print(f"Received DOCX upload: {file.filename}")

    try:
        file_extension = os.path.splitext(file.filename)[1].lower()
        if file_extension != ".docx":
            return JSONResponse(
                status_code=400,
                content={"error": "Unsupported file format. Only .docx files are supported."}
            )

        client_id = str(uuid.uuid4())
        print(f"Generated DOC client ID: {client_id}")

        file_path = os.path.join(UPLOAD_DIR, f"{client_id}{file_extension}")
        print(f"Saving DOCX to: {file_path}")

        with open(file_path, "wb") as buffer:
            content = await file.read()
            print(f"Read {len(content)} bytes from DOCX")
            if not content:
                return JSONResponse(
                    status_code=400,
                    content={"error": "Uploaded DOCX is empty."}
                )
            buffer.write(content)

        doc_text = extract_docx_text(file_path)
        if not doc_text.strip():
            return JSONResponse(
                status_code=400,
                content={"error": "Failed to extract text from the DOCX."}
            )

        system_message = _build_doc_system_message(file.filename, doc_text)

        active_doc_connections[client_id] = {
            "file_path": file_path,
            "filename": file.filename,
            "doc_text": doc_text,
            "message_history": [
                {"role": "system", "content": system_message}
            ]
        }

        save_upload(client_id, "doc", file_path, file.filename)
        save_session(client_id, "doc", active_doc_connections[client_id]["message_history"])

        print("DOCX upload successful")
        return {
            "client_id": client_id,
            "filename": file.filename,
            "text": doc_text
        }
    except Exception as e:
        error_msg = f"Error in upload_doc: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.get("/excel/{client_id}")
async def get_excel_data(client_id: str):
    """Get the current Excel data for a client"""
    if client_id not in active_connections:
        if not _restore_excel_session(client_id):
            return JSONResponse(status_code=404, content={"error": "Client not found"})
    
    excel_utils = active_connections[client_id]["excel_utils"]
    df = excel_utils.get_dataframe()
    
    # Convert DataFrame to JSON-serializable format
    data = []
    for i in range(len(df)):
        row = {}
        for j in range(len(df.columns)):
            cell_value = df.iloc[i, j]
            # Handle non-serializable types
            if pd.isna(cell_value):
                cell_value = None
            elif isinstance(cell_value, pd.Timestamp) or hasattr(cell_value, 'isoformat'):
                cell_value = cell_value.isoformat()
            # Handle NumPy types by converting to Python native types
            elif hasattr(cell_value, 'item'):
                try:
                    cell_value = cell_value.item()
                except (ValueError, TypeError):
                    cell_value = str(cell_value)
            elif not isinstance(cell_value, (str, int, float, bool, type(None))):
                cell_value = str(cell_value)
            row[str(j)] = cell_value
        data.append(row)
    
    # Get metadata
    rows, cols = df.shape
    
    return {
        "data": data,
        "metadata": {
            "rows": rows,
            "columns": cols
        }
    }

@app.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    """WebSocket endpoint for real-time communication with the Excel agent"""
    if client_id not in active_connections:
        if not _restore_excel_session(client_id):
            await websocket.close(code=1000, reason="Client not found")
            return
    
    await manager.connect(websocket, client_id)
    
    try:
        while True:
            # Receive message from client
            data = await websocket.receive_text()
            message_data = json.loads(data)
            if message_data.get("type") == "cancel":
                _mark_canceled(client_id, message_data.get("request_id"))
                await manager.send_message(
                    client_id,
                    json.dumps({
                        "type": "canceled",
                        "request_id": message_data.get("request_id")
                    })
                )
                continue
            user_message = message_data.get("message", "")
            request_id = message_data.get("request_id")
            model_provider = message_data.get("model_provider", "openai")
            model_id = message_data.get("model_id")
            model_params = message_data.get("model_params") or {}
            if model_provider not in {"openai", "bedrock"}:
                model_provider = "openai"
            if model_provider == "bedrock" and model_params.get("bedrock_use_attachment"):
                model_params["file_path"] = active_connections[client_id]["file_path"]
            if _is_canceled(client_id, request_id):
                continue
            
            # Add user message to history
            active_connections[client_id]["message_history"].append({"role": "user", "content": user_message})
            _persist_message_history(client_id, "spreadsheet", active_connections[client_id]["message_history"])
            
            # Process with the agent
            excel_utils = active_connections[client_id]["excel_utils"]
            message_history = active_connections[client_id]["message_history"]
            
            # Call the agent
            try:
                response, excel_modified = excel_agent.call_agent(
                    message_history,
                    excel_utils,
                    model_provider,
                    model_id,
                    model_params
                )

                if _is_canceled(client_id, request_id):
                    canceled_requests.get(client_id, set()).discard(request_id)
                    continue
                
                # Add assistant response to history
                active_connections[client_id]["message_history"].append({"role": "assistant", "content": response})
                _persist_message_history(client_id, "spreadsheet", active_connections[client_id]["message_history"])
                
                # Send response back to client
                try:
                    await manager.send_message(
                        client_id, 
                        json.dumps({
                            "response": response,
                            "excel_modified": excel_modified,
                            "request_id": request_id
                        })
                    )
                except TypeError as e:
                    print(f"JSON serialization error in response: {str(e)}")
                    # Try to serialize with a custom encoder
                    await manager.send_message(
                        client_id, 
                        json.dumps({
                            "response": str(response),
                            "excel_modified": excel_modified,
                            "request_id": request_id
                        })
                    )
                
                # If Excel was modified, notify client to refresh the data
                if excel_modified:
                    # Get updated Excel data
                    df = excel_utils.get_dataframe()
                    
                    # Convert DataFrame to JSON-serializable format
                    data = []
                    for i in range(len(df)):
                        row = {}
                        for j in range(len(df.columns)):
                            cell_value = df.iloc[i, j]
                            # Handle non-serializable types
                            if pd.isna(cell_value):
                                cell_value = None
                            elif isinstance(cell_value, pd.Timestamp) or hasattr(cell_value, 'isoformat'):
                                cell_value = cell_value.isoformat()
                            elif not isinstance(cell_value, (str, int, float, bool, type(None))):
                                cell_value = str(cell_value)
                            row[str(j)] = cell_value
                        data.append(row)
                    
                    # Get metadata
                    rows, cols = df.shape
                    
                    # Send updated data
                    try:
                        await manager.send_message(
                            client_id,
                            json.dumps({
                                "type": "excel_update",
                                "data": data,
                                "metadata": {
                                    "rows": rows,
                                    "columns": cols
                                },
                                "request_id": request_id
                            })
                        )
                    except TypeError as e:
                        print(f"JSON serialization error in excel_update: {str(e)}")
            except Exception as e:
                print(f"Error in agent processing: {str(e)}")
                if not _is_canceled(client_id, request_id):
                    await manager.send_message(
                        client_id,
                        json.dumps({
                            "response": f"Error processing request: {str(e)}",
                            "excel_modified": False,
                            "request_id": request_id
                        })
                    )
    
    except WebSocketDisconnect:
        manager.disconnect(client_id)


@app.websocket("/ws/gemini/{client_id}")
async def gemini_websocket_endpoint(websocket: WebSocket, client_id: str):
    """WebSocket endpoint for real-time communication with the Gemini agent"""
    if client_id not in active_gemini_connections:
        if not _restore_gemini_session(client_id):
            await websocket.close(code=1000, reason="Client not found")
            return

    session_key = _gemini_session_key(client_id)
    await manager.connect(websocket, session_key)

    try:
        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)
            if message_data.get("type") == "cancel":
                _mark_canceled(session_key, message_data.get("request_id"))
                await manager.send_message(
                    session_key,
                    json.dumps({
                        "type": "canceled",
                        "request_id": message_data.get("request_id")
                    })
                )
                continue

            user_message = message_data.get("message", "")
            request_id = message_data.get("request_id")
            model_id = message_data.get("model_id")
            model_params = message_data.get("model_params") or {}
            raw_attachments = message_data.get("image_attachments") or []
            image_attachments = []
            for attachment in raw_attachments:
                if not isinstance(attachment, dict):
                    continue
                mime_type = attachment.get("mime_type")
                data = attachment.get("data")
                if not mime_type or not data:
                    continue
                try:
                    image_bytes = base64.b64decode(data)
                except (ValueError, TypeError) as error:
                    print(f"Failed to decode image attachment: {error}")
                    continue
                image_attachments.append({
                    "mime_type": mime_type,
                    "data": image_bytes
                })
            if _is_canceled(session_key, request_id):
                continue

            active_gemini_connections[client_id]["message_history"].append(
                {"role": "user", "content": user_message}
            )
            _persist_gemini_message_history(client_id, active_gemini_connections[client_id]["message_history"])

            message_history = active_gemini_connections[client_id]["message_history"]
            tool_runner = None
            if active_gemini_connections[client_id].get("type") == "spreadsheet":
                excel_utils = active_gemini_connections[client_id].get("excel_utils")
                if excel_utils:
                    def tool_runner(tool_name, tool_args):
                        try:
                            if tool_name == "read_excel_range":
                                return excel_utils.read_excel_range(tool_args.get("range", ""))
                            if tool_name == "read_cell":
                                return excel_utils.read_cell(
                                    cell=tool_args.get("cell"),
                                    row=tool_args.get("row"),
                                    col=tool_args.get("col")
                                )
                            if tool_name == "read_sheet_metadata":
                                return excel_utils.read_sheet_metadata()
                            if tool_name == "get_column_values":
                                return excel_utils.get_column_values(int(tool_args.get("index", 0)))
                            if tool_name == "filter_rows":
                                return excel_utils.filter_rows(
                                    int(tool_args.get("col_index", 0)),
                                    str(tool_args.get("value", ""))
                                )
                            if tool_name == "get_last_filled_row_index":
                                return str(excel_utils.get_last_filled_row_index())
                            return f"Unknown tool: {tool_name}"
                        except Exception as error:
                            return f"Tool error ({tool_name}): {error}"

            try:
                response = gemini_agent.call_agent(
                    message_history,
                    model_id,
                    model_params,
                    image_attachments,
                    tool_runner=tool_runner
                )

                if _is_canceled(session_key, request_id):
                    canceled_requests.get(session_key, set()).discard(request_id)
                    continue

                active_gemini_connections[client_id]["message_history"].append(
                    {"role": "assistant", "content": response}
                )
                _persist_gemini_message_history(client_id, active_gemini_connections[client_id]["message_history"])

                await manager.send_message(
                    session_key,
                    json.dumps({
                        "response": response,
                        "request_id": request_id
                    })
                )
            except Exception as e:
                print(f"Error in Gemini agent processing: {str(e)}")
                if not _is_canceled(session_key, request_id):
                    await manager.send_message(
                        session_key,
                        json.dumps({
                            "response": f"Error processing request: {str(e)}",
                            "request_id": request_id
                        })
                    )

    except WebSocketDisconnect:
        manager.disconnect(session_key)
    except Exception as e:
        print(f"Error: {str(e)}")
        manager.disconnect(session_key)

@app.get("/pdf/{client_id}/file")
async def get_pdf_file(client_id: str):
    upload = get_upload(client_id)
    if not upload or upload["type"] != "pdf":
        return JSONResponse(status_code=404, content={"error": "PDF client not found"})
    file_path = upload["file_path"]
    if not os.path.exists(file_path):
        return JSONResponse(status_code=404, content={"error": "PDF file not found"})
    return FileResponse(file_path, media_type="application/pdf")


@app.get("/doc/{client_id}/text")
async def get_doc_text(client_id: str):
    if client_id not in active_doc_connections:
        if not _restore_doc_session(client_id):
            return JSONResponse(status_code=404, content={"error": "DOC client not found"})
    return {"text": active_doc_connections[client_id]["doc_text"]}


@app.delete("/uploads/{client_id}")
async def delete_upload_endpoint(client_id: str):
    upload = get_upload(client_id)
    if not upload:
        return JSONResponse(status_code=404, content={"error": "Upload not found"})

    file_path = upload["file_path"]
    if client_id in active_connections:
        del active_connections[client_id]
    if client_id in active_pdf_connections:
        del active_pdf_connections[client_id]
    if client_id in active_doc_connections:
        del active_doc_connections[client_id]
    manager.disconnect(client_id)
    if client_id in active_gemini_connections:
        del active_gemini_connections[client_id]
    manager.disconnect(_gemini_session_key(client_id))

    if file_path and os.path.exists(file_path):
        try:
            os.remove(file_path)
        except OSError as e:
            return JSONResponse(
                status_code=500,
                content={"error": f"Failed to delete file: {str(e)}"}
            )

    delete_upload(client_id)
    delete_session(_gemini_session_key(client_id))
    return {"status": "deleted", "client_id": client_id}

@app.websocket("/ws/pdf/{client_id}")
async def pdf_websocket_endpoint(websocket: WebSocket, client_id: str):
    """WebSocket endpoint for real-time communication with the PDF agent"""
    if client_id not in active_pdf_connections:
        if not _restore_pdf_session(client_id):
            await websocket.close(code=1000, reason="Client not found")
            return

    await manager.connect(websocket, client_id)

    try:
        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)
            if message_data.get("type") == "cancel":
                _mark_canceled(client_id, message_data.get("request_id"))
                await manager.send_message(
                    client_id,
                    json.dumps({
                        "type": "canceled",
                        "request_id": message_data.get("request_id")
                    })
                )
                continue
            user_message = message_data.get("message", "")
            request_id = message_data.get("request_id")
            model_provider = message_data.get("model_provider", "openai")
            model_id = message_data.get("model_id")
            model_params = message_data.get("model_params") or {}
            if model_provider not in {"openai", "bedrock"}:
                model_provider = "openai"
            if _is_canceled(client_id, request_id):
                continue

            active_pdf_connections[client_id]["message_history"].append(
                {"role": "user", "content": user_message}
            )
            _persist_message_history(client_id, "pdf", active_pdf_connections[client_id]["message_history"])

            message_history = active_pdf_connections[client_id]["message_history"]

            try:
                response = pdf_agent.call_agent(
                    message_history,
                    model_provider,
                    model_id,
                    model_params
                )

                if _is_canceled(client_id, request_id):
                    canceled_requests.get(client_id, set()).discard(request_id)
                    continue

                active_pdf_connections[client_id]["message_history"].append(
                    {"role": "assistant", "content": response}
                )
                _persist_message_history(client_id, "pdf", active_pdf_connections[client_id]["message_history"])

                await manager.send_message(
                    client_id,
                    json.dumps({
                        "response": response,
                        "request_id": request_id
                    })
                )
            except Exception as e:
                print(f"Error in PDF agent processing: {str(e)}")
                if not _is_canceled(client_id, request_id):
                    await manager.send_message(
                        client_id,
                        json.dumps({
                            "response": f"Error processing request: {str(e)}",
                            "request_id": request_id
                        })
                    )

    except WebSocketDisconnect:
        manager.disconnect(client_id)
    except Exception as e:
        print(f"Error: {str(e)}")
        manager.disconnect(client_id)
    except Exception as e:
        print(f"Error: {str(e)}")
        manager.disconnect(client_id)


@app.websocket("/ws/doc/{client_id}")
async def doc_websocket_endpoint(websocket: WebSocket, client_id: str):
    """WebSocket endpoint for real-time communication with the DOCX agent"""
    if client_id not in active_doc_connections:
        if not _restore_doc_session(client_id):
            await websocket.close(code=1000, reason="Client not found")
            return

    await manager.connect(websocket, client_id)

    try:
        while True:
            data = await websocket.receive_text()
            message_data = json.loads(data)
            if message_data.get("type") == "cancel":
                _mark_canceled(client_id, message_data.get("request_id"))
                await manager.send_message(
                    client_id,
                    json.dumps({
                        "type": "canceled",
                        "request_id": message_data.get("request_id")
                    })
                )
                continue
            user_message = message_data.get("message", "")
            request_id = message_data.get("request_id")
            model_provider = message_data.get("model_provider", "openai")
            model_id = message_data.get("model_id")
            model_params = message_data.get("model_params") or {}
            if model_provider not in {"openai"}:
                model_provider = "openai"
            if _is_canceled(client_id, request_id):
                continue

            active_doc_connections[client_id]["message_history"].append(
                {"role": "user", "content": user_message}
            )
            _persist_message_history(client_id, "doc", active_doc_connections[client_id]["message_history"])

            message_history = active_doc_connections[client_id]["message_history"]

            try:
                response = doc_agent.call_agent(
                    message_history,
                    model_provider,
                    model_id,
                    model_params
                )

                if _is_canceled(client_id, request_id):
                    canceled_requests.get(client_id, set()).discard(request_id)
                    continue

                active_doc_connections[client_id]["message_history"].append(
                    {"role": "assistant", "content": response}
                )
                _persist_message_history(client_id, "doc", active_doc_connections[client_id]["message_history"])

                await manager.send_message(
                    client_id,
                    json.dumps({
                        "response": response,
                        "request_id": request_id
                    })
                )
            except Exception as e:
                print(f"Error in DOCX agent processing: {str(e)}")
                if not _is_canceled(client_id, request_id):
                    await manager.send_message(
                        client_id,
                        json.dumps({
                            "response": f"Error processing request: {str(e)}",
                            "request_id": request_id
                        })
                    )

    except WebSocketDisconnect:
        manager.disconnect(client_id)
    except Exception as e:
        print(f"Error: {str(e)}")
        manager.disconnect(client_id)
    except Exception as e:
        print(f"Error: {str(e)}")
        manager.disconnect(client_id)

@app.on_event("startup")
async def startup_event():
    # Create uploads directory if it doesn't exist
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    init_db()
    _bootstrap_upload_records()

if __name__ == "__main__":
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
